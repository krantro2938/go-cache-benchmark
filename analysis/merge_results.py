#!/usr/bin/env python3
"""
Merge benchmark results from multiple runs into a single Excel file.
Each metric type becomes a separate sheet with all runs combined.
"""

import os
import csv
import glob
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

RESULTS_DIR = "../../res"
OUTPUT_FILE = "merged_results.xlsx"

METRIC_FILES = [
    "latency.csv",
    "throughput.csv", 
    "hitratio.csv",
    "evictions.csv",
    "memory.csv",
    "gc.csv"
]


def read_csv(filepath):
    """Read CSV file and return header and rows."""
    with open(filepath, 'r') as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)
    return header, rows


def merge_metric(metric_file):
    """Merge a specific metric from all runs."""
    all_rows = []
    header = None
    
    # Find all run directories
    run_dirs = sorted(glob.glob(os.path.join(RESULTS_DIR, "run_*")))
    
    for run_dir in run_dirs:
        run_name = os.path.basename(run_dir)
        filepath = os.path.join(run_dir, metric_file)
        
        if not os.path.exists(filepath):
            print(f"Warning: {filepath} not found, skipping...")
            continue
        
        h, rows = read_csv(filepath)
        
        if header is None:
            header = ['run'] + h
        
        # Add run identifier to each row
        for row in rows:
            all_rows.append([run_name] + row)
    
    return header, all_rows


def write_excel(data_dict, output_file):
    """Write all metrics to Excel with each metric as a sheet."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for metric_name, (header, rows) in data_dict.items():
        # Create sheet name from metric filename (without .csv)
        sheet_name = metric_name.replace('.csv', '')
        ws = wb.create_sheet(title=sheet_name)
        
        # Write header
        for col, value in enumerate(header, 1):
            ws.cell(row=1, column=col, value=value)
        
        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(header) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
    
    wb.save(output_file)
    print(f"Saved: {output_file}")


def main():
    print("Merging benchmark results...")
    print(f"Results directory: {RESULTS_DIR}")
    print()
    
    data_dict = {}
    
    for metric_file in METRIC_FILES:
        print(f"Processing: {metric_file}")
        header, rows = merge_metric(metric_file)
        data_dict[metric_file] = (header, rows)
        print(f"  -> {len(rows)} rows")
    
    write_excel(data_dict, OUTPUT_FILE)
    print()
    print("Done!")


if __name__ == "__main__":
    main()
